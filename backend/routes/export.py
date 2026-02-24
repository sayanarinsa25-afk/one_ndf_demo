from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import FileResponse
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from datetime import datetime
import tempfile
import os

router = APIRouter()


# ================= DATA SOURCE =================
def get_dashboard_data():
    """
    Replace this with DB/service layer in production.
    """
    return {
        "total_leads": 13,
        "applications": 8,
        "approval_rate": "25%",
        "avg_risk_score": 64,
        "pipeline_value": "₹5.5Cr",
        "approved": 2,
        "rejected": 2,
    }


# ================= PDF GENERATOR =================
def generate_pdf(file_path: str, data: dict):
    """Create professional banking-style MIS PDF."""

    # Unicode font for ₹ symbol
    pdfmetrics.registerFont(UnicodeCIDFont("HYSMyeongJo-Medium"))

    doc = SimpleDocTemplate(file_path)
    styles = getSampleStyleSheet()
    elements = []

    # Title + timestamp
    elements.append(Paragraph("Finance AI — Loan MIS Report", styles["Title"]))
    elements.append(Spacer(1, 12))
    elements.append(
        Paragraph(
            f"Generated on: {datetime.now().strftime('%d %b %Y, %I:%M %p')}",
            styles["Normal"],
        )
    )
    elements.append(Spacer(1, 20))

    # Table
    table_data = [
        ["Metric", "Value"],
        ["Total Leads", data["total_leads"]],
        ["Applications", data["applications"]],
        ["Approval Rate", data["approval_rate"]],
        ["Average Risk Score", data["avg_risk_score"]],
        ["Pipeline Value", data["pipeline_value"]],
        ["Loans Approved", data["approved"]],
        ["Loans Rejected", data["rejected"]],
    ]

    table = Table(table_data, colWidths=[240, 160])

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2563eb")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, -1), "HYSMyeongJo-Medium"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTSIZE", (0, 0), (-1, 0), 12),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ]
        )
    )

    elements.append(table)
    doc.build(elements)


# ================= EXCEL GENERATOR =================
def generate_excel(file_path: str, data: dict):
    """Create styled Excel MIS."""

    wb = Workbook()
    ws = wb.active
    ws.title = "Loan MIS"

    # Title row
    ws.append(["Finance AI — Loan MIS Report"])
    ws.merge_cells("A1:B1")
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["Metric", "Value"])

    ws["A3"].font = Font(bold=True)
    ws["B3"].font = Font(bold=True)

    rows = [
        ("Total Leads", data["total_leads"]),
        ("Applications", data["applications"]),
        ("Approval Rate", data["approval_rate"]),
        ("Average Risk Score", data["avg_risk_score"]),
        ("Pipeline Value", data["pipeline_value"]),
        ("Loans Approved", data["approved"]),
        ("Loans Rejected", data["rejected"]),
    ]

    for r in rows:
        ws.append(r)

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20

    wb.save(file_path)


# ================= EXPORT ENDPOINT =================
@router.get("/export-mis")
def export_mis(type: str = Query(..., pattern="^(pdf|excel)$")):
    """
    Download MIS report in PDF or Excel.
    Uses temporary files (safe for production).
    """

    data = get_dashboard_data()

    # create temp file
    suffix = ".pdf" if type == "pdf" else ".xlsx"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        file_path = tmp.name

    try:
        if type == "pdf":
            generate_pdf(file_path, data)
            filename = "Loan_MIS_Report.pdf"

        elif type == "excel":
            generate_excel(file_path, data)
            filename = "Loan_MIS_Report.xlsx"

        else:
            raise HTTPException(status_code=400, detail="Invalid export type")

        return FileResponse(
            path=file_path,
            filename=filename,
            media_type="application/octet-stream",
        )

    finally:
        # optional cleanup can be scheduled later via background task
        pass
